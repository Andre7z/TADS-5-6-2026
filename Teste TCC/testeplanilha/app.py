from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Alunos"

# Cabeçalho
headers = ["Nome", "Idade", "Faixa"]
ws.append(headers)

# Dados
dados = [
    ["Arthur", 11, "Cinza"],
    ["João", 12, "Amarela"],
    ["Maria", 8, "Azul"]
]

for linha in dados:
    ws.append(linha)

# ===== ESTILO DO CABEÇALHO =====
for col in range(1, 4):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True, color="FFFFFF")  
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # azul
    cell.alignment = Alignment(horizontal="center")

# ===== BORDAS =====
borda = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=3):
    for cell in row:
        cell.border = borda
        cell.alignment = Alignment(horizontal="center")

# ===== AJUSTAR LARGURA DAS COLUNAS =====
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 15

# Salvar
wb.save("alunos_formatado.xlsx")

print("Planilha estilizada criada!")